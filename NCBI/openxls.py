from openpyxl import load_workbook
import os

def getOpenxls():
    f = os.getcwd()
    burl = 'https://www.ncbi.nlm.nih.gov/pubmed/'
    f_a = f + '\\pmid.xlsx'
    wb = load_workbook(f_a)
    #sheet1 = wb.get_sheet_by_name('condition')
    sheet2 = wb.get_sheet_by_name('snp_score')
    #xlpath = sheet1['D4'].value
    urlpath = []
    for r in range(1,sheet2.max_row):
        url = sheet2['C'][r].value
        if url:
            urlpath.append(burl + str(url))

    return set(urlpath)
